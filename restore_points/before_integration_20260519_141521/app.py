import base64
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from zipfile import BadZipFile

import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook


APP_TITLE = "Disponibilidade Comercial"
BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = Path(
    r"C:\Users\rafael.germano\Documents\Rafael Germano\Codex\2026-05-19\app disponiblidade\disponibilidade.xlsx"
)
LOGO_PATH = BASE_DIR / "assets" / "trebeschi_logo.png"
AUTO_REFRESH_SECONDS = 120

CULTURE_ICONS = {
    "Tomate": "🍅",
    "Cebola": "🧅",
    "Batata Doce Rosada": "🍠",
    "Alho": "🧄",
}


@dataclass(frozen=True)
class LoadingLocation:
    name: str
    cultures: list[str]
    excel_locations: list[str]


class AppCatalog:
    def __init__(self) -> None:
        self.culture_order = ["Tomate", "Alho", "Cebola", "Batata Doce Rosada"]
        self.locations = {
            "Brasil": LoadingLocation(
                name="Brasil",
                cultures=self.culture_order,
                excel_locations=[],
            ),
            "Araguari": LoadingLocation(
                name="Araguari",
                cultures=["Tomate", "Alho", "Cebola"],
                excel_locations=["Araguari", "Trebeschi", "Mandaguari"],
            ),
            "Anápolis": LoadingLocation(
                name="Anápolis",
                cultures=["Tomate", "Batata Doce Rosada"],
                excel_locations=["Anápolis"],
            ),
            "Cascavel": LoadingLocation(
                name="Cascavel",
                cultures=["Tomate"],
                excel_locations=["Cascavel"],
            ),
        }

    def location_names(self) -> list[str]:
        return list(self.locations.keys())

    def cultures_for_location(self, location: str, items: list[dict]) -> list[str]:
        if location == "Brasil":
            present = {item["cultura"] for item in items}
            return [culture for culture in self.culture_order if culture in present]

        selected = self.locations.get(location)
        return selected.cultures if selected else []

    def excel_locations_for_filter(self, location: str) -> list[str]:
        selected = self.locations.get(location)
        return selected.excel_locations if selected else [location]


class ExcelAvailabilityRepository:
    def __init__(self, excel_path: Path, culture_names: list[str]) -> None:
        self.excel_path = excel_path
        self.culture_names = set(culture_names)

    def list_items(self) -> list[dict]:
        if not self.excel_path.exists():
            st.warning(f"Planilha nao encontrada: {self.excel_path}")
            return []

        try:
            workbook = load_workbook(self.excel_path, data_only=True, read_only=True)
        except (PermissionError, OSError, BadZipFile):
            st.warning(
                "A planilha esta sendo salva pelo Excel. Aguarde a proxima atualizacao automatica."
            )
            return []

        try:
            items = []
            for sheet_name in workbook.sheetnames:
                if sheet_name in self.culture_names:
                    items.extend(self._read_culture_sheet(workbook[sheet_name], sheet_name))
            return items
        finally:
            workbook.close()

    def _read_culture_sheet(self, sheet, culture: str) -> list[dict]:
        matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
        items = []
        next_id = 1

        for row_index, row in enumerate(matrix):
            for col_index, value in enumerate(row):
                if normalize_text(value) != "disponibilidade":
                    continue

                block_items = self._read_block(
                    matrix=matrix,
                    culture=culture,
                    start_row=row_index,
                    start_col=col_index,
                    first_id=next_id,
                )
                items.extend(block_items)
                next_id += len(block_items)

        return items

    def _read_block(
        self,
        matrix: list[list],
        culture: str,
        start_row: int,
        start_col: int,
        first_id: int,
    ) -> list[dict]:
        location = cell_text(matrix, start_row + 3, start_col)
        packing = cell_text(matrix, start_row + 3, start_col + 2)
        updated_at = format_date(cell_value(matrix, start_row + 1, start_col))

        if not location or not packing:
            return []

        items = []
        row_index = start_row + 4
        while row_index < len(matrix):
            product = cell_text(matrix, row_index, start_col)
            if not product or normalize_text(product) == "total":
                break

            quantity = parse_number(cell_value(matrix, row_index, start_col + 1))
            items.append(
                {
                    "id": first_id + len(items),
                    "local_carregamento": location,
                    "cultura": culture,
                    "packing": packing,
                    "produto": product,
                    "ordem_packing": start_row * 1000 + start_col,
                    "ordem_produto": row_index - (start_row + 4),
                    "quantidade": quantity,
                    "atualizado_em": updated_at,
                }
            )
            row_index += 1

        return items


class PageStyle:
    @staticmethod
    def apply() -> None:
        st.set_page_config(page_title=APP_TITLE, page_icon="📦", layout="centered")
        st.markdown(
            """
            <style>
                .block-container {
                    max-width: 980px;
                    padding-top: 1rem;
                    padding-bottom: 3rem;
                }
                div[data-testid="stMetric"] {
                    border: 1px solid #d7dde5;
                    border-radius: 8px;
                    padding: 0.7rem 0.85rem;
                    background: #ffffff;
                }
                .app-header {
                    display: flex;
                    align-items: center;
                    gap: 1.1rem;
                    margin-bottom: 0.35rem;
                }
                .app-logo {
                    width: 124px;
                    max-width: 32vw;
                    height: auto;
                    object-fit: contain;
                    display: block;
                }
                .app-title {
                    font-size: 2rem;
                    line-height: 1.1;
                    font-weight: 750;
                    color: #1f2937;
                }
                .culture-title {
                    margin-top: 1.4rem;
                    margin-bottom: 0.45rem;
                    font-size: 1.2rem;
                    font-weight: 750;
                    color: #1f2937;
                }
                .stDataFrame {
                    border: 1px solid #d7dde5;
                    border-radius: 8px;
                    overflow: hidden;
                }
                .availability-table {
                    width: 100%;
                    border-collapse: collapse;
                    font-size: 0.92rem;
                    margin-bottom: 1rem;
                    background: #ffffff;
                    border: 1px solid #d7dde5;
                    border-radius: 8px;
                    overflow: hidden;
                }
                .availability-table th {
                    background: #1f6f5b;
                    color: #ffffff;
                    font-weight: 700;
                    text-align: center;
                    padding: 0.45rem 0.5rem;
                    border: 1px solid #d7dde5;
                }
                .availability-table td {
                    text-align: center;
                    padding: 0.42rem 0.5rem;
                    border: 1px solid #d7dde5;
                }
                .availability-table tbody tr:nth-child(odd) td {
                    background: #edf5e8;
                }
                .availability-table tbody tr.total-row td {
                    background: #2e75b6;
                    color: #ffffff;
                    font-weight: 700;
                }
                .table-subtitle {
                    text-align: center;
                    font-weight: 750;
                    color: #1f2937;
                    margin: 0.25rem 0 0.4rem;
                }
                .app-footer {
                    margin-top: 2rem;
                    padding-top: 0.85rem;
                    border-top: 1px solid #d7dde5;
                    color: #526070;
                    text-align: center;
                    font-size: 0.85rem;
                }
                @media (max-width: 640px) {
                    .app-header {
                        align-items: flex-start;
                    }
                    .app-logo {
                        width: 92px;
                    }
                    .app-title {
                        font-size: 1.55rem;
                    }
                }
            </style>
            """,
            unsafe_allow_html=True,
        )


class AvailabilityApp:
    def __init__(self) -> None:
        self.catalog = AppCatalog()
        self.repository = ExcelAvailabilityRepository(
            excel_path=EXCEL_PATH,
            culture_names=self.catalog.culture_order,
        )

    def run(self) -> None:
        PageStyle.apply()
        self.render_header()
        st.caption(
            "Consulta de produtos disponiveis por local de carregamento, atualizada via Excel."
        )
        self.render_update_status()
        self.enable_auto_refresh()
        self.render_consultation(self.repository.list_items())
        self.render_footer()

    def render_header(self) -> None:
        logo_html = ""
        if LOGO_PATH.exists():
            logo_data = base64.b64encode(LOGO_PATH.read_bytes()).decode("ascii")
            logo_html = (
                f'<img class="app-logo" src="data:image/png;base64,{logo_data}" '
                'alt="Trebeschi">'
            )

        st.markdown(
            f"""
            <div class="app-header">
                {logo_html}
                <div class="app-title">{APP_TITLE}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    def render_update_status(self) -> None:
        if not EXCEL_PATH.exists():
            st.warning(f"Planilha disponibilidade.xlsx nao encontrada em: {EXCEL_PATH}")
            return

        modified_at = datetime.fromtimestamp(EXCEL_PATH.stat().st_mtime)
        st.caption(
            f"Ultima atualizacao da planilha: {modified_at.strftime('%d/%m/%Y %H:%M:%S')} "
            "| Atualizacao automatica a cada 2 minutos"
        )
        st.caption(f"Arquivo lido: {EXCEL_PATH}")

    @staticmethod
    def enable_auto_refresh() -> None:
        components.html(
            f"""
            <script>
                window.parent.document.documentElement.setAttribute("lang", "pt-BR");
                window.parent.document.body.setAttribute("translate", "no");
                setTimeout(function() {{
                    window.parent.location.reload();
                }}, {AUTO_REFRESH_SECONDS * 1000});
            </script>
            """,
            height=0,
        )

    def render_consultation(self, items: list[dict]) -> None:
        st.subheader("Disponibilidade por local")

        if not items:
            st.info("Nenhum produto encontrado na planilha.")
            return

        selected_location = st.selectbox(
            "Local de carregamento",
            self.catalog.location_names(),
            index=0,
        )
        available_cultures = self.catalog.cultures_for_location(selected_location, items)
        selected_culture_label = st.selectbox(
            "Cultura",
            ["Todas"] + [culture_label(culture) for culture in available_cultures],
        )
        selected_culture = culture_from_label(selected_culture_label)

        search = st.text_input(
            "Buscar produto ou packing",
            placeholder="Ex: AA Misto, Classe 5, LV",
        )

        filtered = self.filter_items(
            items=items,
            location=selected_location,
            culture=selected_culture,
            search=search,
        )
        filtered = [item for item in filtered if item["quantidade"] > 0]

        col1, col2 = st.columns(2)
        col1.metric("Itens", len(filtered))
        col2.metric("Qtd. disponivel", format_quantity(sum_quantities(filtered)) or "0")

        if not filtered:
            st.warning("Nenhum item encontrado com os filtros atuais.")
            return

        cultures_to_render = (
            [selected_culture]
            if selected_culture != "Todas"
            else available_cultures
        )
        for culture in cultures_to_render:
            culture_items = [item for item in filtered if item["cultura"] == culture]
            if culture_items:
                self.render_culture_table(culture, culture_items)

    def filter_items(
        self,
        items: list[dict],
        location: str,
        culture: str,
        search: str,
    ) -> list[dict]:
        filtered = list(items)

        if location != "Brasil":
            allowed_locations = self.catalog.excel_locations_for_filter(location)
            filtered = [
                item
                for item in filtered
                if item["local_carregamento"] in allowed_locations
            ]

        if culture != "Todas":
            filtered = [item for item in filtered if item["cultura"] == culture]

        if search:
            term = normalize_text(search)
            filtered = [
                item
                for item in filtered
                if term in normalize_text(item["produto"])
                or term in normalize_text(item["packing"])
                or term in normalize_text(item["local_carregamento"])
            ]

        return filtered

    def render_culture_table(self, culture: str, items: list[dict]) -> None:
        st.markdown(
            f'<div class="culture-title">{culture_label(culture)}</div>',
            unsafe_allow_html=True,
        )

        split_packings = unique_by_order(items, "packing", "ordem_packing")
        if len(split_packings) > 1:
            columns = st.columns(len(split_packings))
            for column, packing in zip(columns, split_packings):
                packing_items = [item for item in items if item["packing"] == packing]
                with column:
                    self.render_quantity_table(packing_title(culture, packing), packing_items)
            return

        self.render_quantity_table("", items)

    def render_quantity_table(self, title: str, items: list[dict]) -> None:
        if not items:
            return

        if title:
            st.markdown(f'<div class="table-subtitle">{title}</div>', unsafe_allow_html=True)

        columns = unique_quantity_columns(items)
        products = unique_by_order(items, "produto", "ordem_produto")

        rows = []
        for product in products:
            row = {"Produto": product}
            row_total = 0
            for column in columns:
                quantity = sum(
                    item["quantidade"]
                    for item in items
                    if item["produto"] == product
                    and quantity_column_name(item, items) == column
                )
                row[column] = quantity
                row_total += quantity
            if row_total > 0:
                row["Total"] = row_total
                rows.append(row)

        total_row = {"Produto": "TOTAL"}
        for column in columns:
            total_row[column] = sum(
                item["quantidade"]
                for item in items
                if quantity_column_name(item, items) == column
            )
        total_row["Total"] = sum_quantities(items)
        rows.append(total_row)

        st.markdown(render_html_table(rows), unsafe_allow_html=True)

    @staticmethod
    def render_footer() -> None:
        st.markdown(
            """
            <div class="app-footer">
                Desenvolvido por Rafael Germano.<br>
                Para uso interno da Trebeschi.
            </div>
            """,
            unsafe_allow_html=True,
        )


def cell_value(matrix: list[list], row: int, col: int):
    if row >= len(matrix) or col >= len(matrix[row]):
        return None
    return matrix[row][col]


def cell_text(matrix: list[list], row: int, col: int) -> str:
    value = cell_value(matrix, row, col)
    return "" if value is None else str(value).strip()


def parse_number(value) -> float:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(".", "").replace(",", "."))
    except ValueError:
        return 0


def format_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if value:
        return str(value)
    return datetime.now().strftime("%d/%m/%Y")


def normalize_text(value) -> str:
    text = "" if value is None else str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(char for char in text if not unicodedata.combining(char))


def culture_label(culture: str) -> str:
    if culture == "Todas":
        return culture
    return f"{CULTURE_ICONS.get(culture, '•')} {culture}"


def culture_from_label(label: str) -> str:
    if label == "Todas":
        return "Todas"
    return label.split(" ", 1)[1] if " " in label else label


def packing_title(culture: str, packing: str) -> str:
    tomato_titles = {
        "lv": "Longa Vida (LV)",
        "it": "Italiano (IT)",
    }
    if culture == "Tomate":
        return tomato_titles.get(normalize_text(packing), packing)
    return packing


def quantity_column_name(item: dict, items: list[dict]) -> str:
    locations = {row["local_carregamento"] for row in items}
    if len(locations) == 1:
        return item["packing"]
    return f"{item['local_carregamento']} - {item['packing']}"


def format_quantity(value: float) -> str:
    if not value:
        return ""
    return f"{value:,.0f}".replace(",", ".")


def render_html_table(rows: list[dict]) -> str:
    if not rows:
        return ""

    headers = list(rows[0].keys())
    header_html = "".join(
        f'<th translate="no">{escape_html(header)}</th>' for header in headers
    )
    body_html = []

    for row in rows:
        row_class = ' class="total-row"' if row.get("Produto") == "TOTAL" else ""
        cells = []
        for header in headers:
            value = row.get(header, "")
            if header != "Produto":
                value = format_quantity(value)
            cells.append(f'<td translate="no">{escape_html(value)}</td>')
        body_html.append(f"<tr{row_class}>{''.join(cells)}</tr>")

    return (
        '<table class="availability-table" translate="no">'
        f"<thead><tr>{header_html}</tr></thead>"
        f"<tbody>{''.join(body_html)}</tbody>"
        "</table>"
    )


def escape_html(value) -> str:
    text = "" if value is None else str(value)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )


def sum_quantities(items: list[dict]) -> float:
    return sum(item["quantidade"] for item in items)


def unique_in_order(values) -> list:
    seen = set()
    result = []
    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result


def unique_by_order(items: list[dict], value_key: str, order_key: str) -> list:
    ordered_values = {}
    for item in items:
        value = item[value_key]
        order = item.get(order_key, 0)
        if value not in ordered_values or order < ordered_values[value]:
            ordered_values[value] = order
    return [
        value
        for value, _ in sorted(
            ordered_values.items(),
            key=lambda pair: (pair[1], str(pair[0])),
        )
    ]


def unique_quantity_columns(items: list[dict]) -> list:
    ordered_columns = {}
    for item in items:
        column = quantity_column_name(item, items)
        order = item.get("ordem_packing", 0)
        if column not in ordered_columns or order < ordered_columns[column]:
            ordered_columns[column] = order
    return [
        column
        for column, _ in sorted(
            ordered_columns.items(),
            key=lambda pair: (pair[1], str(pair[0])),
        )
    ]


if __name__ == "__main__":
    AvailabilityApp().run()
