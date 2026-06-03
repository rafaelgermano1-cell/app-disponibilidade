import base64
import subprocess
import tempfile
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from io import BytesIO
from math import ceil
from pathlib import Path
from zipfile import BadZipFile

import pytz
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook
from PIL import Image as PILImage


# ============================================================================
# Configuracoes principais
# ============================================================================

APP_TITLE = "Trebeschi Comercial"
BASE_DIR = Path(__file__).resolve().parent

EXCEL_PATH = BASE_DIR / "disponibilidade.xlsx"

LOGO_PATH = BASE_DIR / "assets" / "trebeschi_logo.png"

AUTO_REFRESH_SECONDS = 600
BRASILIA_TZ = pytz.timezone("America/Sao_Paulo")
QUALITY_STANDARDS = ("Padrão A", "Padrão B")
PRICE_SHEET_NAME = "Precos"
GIT_SPREADSHEET_PATH = EXCEL_PATH.name
GIT_SYNC_LOCK_PATH = Path(tempfile.gettempdir()) / "trebeschi_disponibilidade_git_sync.lock"
GIT_SYNC_LOCK_MAX_AGE_SECONDS = 300
GIT_COMMIT_MESSAGE = "Atualiza planilha de disponibilidade"

CULTURE_ICONS = {
    "Tomate": "🍅",
    "Cebola": "🧅",
    "Batata Doce": "🍠",
    "Alho": "🧄",
}

EXCEL_SHEET_ALIASES = {
    "Batata Doce Rosada": "Batata Doce",
}

TAXES = {
    "Funrural (1,5%)": 0.015,
    "Pessoa Jurídica (3,28%)": 0.0328,
    "Pernambuco (21%)": 0.21,
}

PACKAGE_COSTS = {
    "Papelão": 8.00,
    "HB": 4.50,
    "IFCO": 4.50,
    "Plástica": 1.00,
    "SC": 0.00,
}


# ============================================================================
# Catalogo de negocio
# ============================================================================

@dataclass(frozen=True)
class LoadingLocation:
    """Representa um filtro de local no app e os locais correspondentes no Excel."""

    name: str
    cultures: list[str]
    excel_locations: list[str]


class AppCatalog:
    """Centraliza a ordem de culturas e a relacao entre filtros e planilha."""

    def __init__(self) -> None:
        self.culture_order = ["Tomate", "Alho", "Cebola", "Batata Doce"]
        self.locations = {
            "Brasil": LoadingLocation(
                name="Brasil",
                cultures=self.culture_order,
                excel_locations=[],
            ),
            "Araguari": LoadingLocation(
                name="Araguari",
                cultures=["Tomate", "Alho", "Cebola"],
                excel_locations=["Araguari", "Trebeschi", "Mandaguari"],
            ),
            "Anápolis": LoadingLocation(
                name="Anápolis",
                cultures=["Tomate", "Batata Doce"],
                excel_locations=["Anápolis"],
            ),
            "Cascavel": LoadingLocation(
                name="Cascavel",
                cultures=["Tomate"],
                excel_locations=["Cascavel"],
            ),
        }

    def location_names(self) -> list[str]:
        return list(self.locations.keys())

    def cultures_for_location(self, location: str, items: list[dict]) -> list[str]:
        if location == "Brasil":
            present = {item["cultura"] for item in items}
            return [culture for culture in self.culture_order if culture in present]

        selected = self.locations.get(location)
        return selected.cultures if selected else []

    def excel_locations_for_filter(self, location: str) -> list[str]:
        selected = self.locations.get(location)
        return selected.excel_locations if selected else [location]


# ============================================================================
# Leitura da disponibilidade
# ============================================================================

class ExcelAvailabilityRepository:
    """Le a planilha disponibilidade.xlsx no formato de blocos por cultura/packing."""

    def __init__(self, excel_path: Path, culture_names: list[str]) -> None:
        self.excel_path = excel_path
        self.culture_names = set(culture_names)

    def list_items(self) -> list[dict]:
        if not self.excel_path.exists():
            st.warning(f"Planilha nao encontrada: {self.excel_path}")
            return []

        try:
            workbook = load_workbook(self.excel_path, data_only=True, read_only=True)
        except (PermissionError, OSError, BadZipFile):
            st.warning(
                "A planilha esta sendo salva pelo Excel. Aguarde a proxima atualizacao automatica."
            )
            return []

        try:
            items = []
            prices = self._read_prices(workbook)
            for sheet_name in workbook.sheetnames:
                culture_name = EXCEL_SHEET_ALIASES.get(sheet_name, sheet_name)
                if culture_name in self.culture_names:
                    items.extend(
                        self._read_culture_sheet(
                            workbook[sheet_name],
                            culture_name,
                            prices,
                        )
                    )
            return items
        finally:
            workbook.close()

    def _read_prices(self, workbook) -> dict[tuple[str, str, str, str], float]:
        if PRICE_SHEET_NAME not in workbook.sheetnames:
            return {}

        sheet = workbook[PRICE_SHEET_NAME]
        prices = {}
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return prices

        headers = [normalize_text(value) for value in rows[0]]
        try:
            culture_col = headers.index("cultura")
            packing_col = headers.index("packing")
            product_col = headers.index("produto")
            price_col = headers.index("preco")
        except ValueError:
            return prices
        quality_col = headers.index("qualidade") if "qualidade" in headers else None

        for row in rows[1:]:
            culture = EXCEL_SHEET_ALIASES.get(cell_from_row(row, culture_col), cell_from_row(row, culture_col))
            packing = cell_from_row(row, packing_col)
            product = cell_from_row(row, product_col)
            quality = cell_from_row(row, quality_col) if quality_col is not None else ""
            price = parse_number(cell_from_row(row, price_col))
            if culture and packing and product and price > 0:
                prices[price_key(culture, packing, product, quality)] = price
        return prices

    def _read_culture_sheet(self, sheet, culture: str, prices: dict[tuple[str, str, str, str], float]) -> list[dict]:
        matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
        items = []
        next_id = 1

        for row_index, row in enumerate(matrix):
            for col_index, value in enumerate(row):
                if normalize_text(value) != "disponibilidade":
                    continue

                block_items = self._read_block(
                    matrix=matrix,
                    culture=culture,
                    prices=prices,
                    start_row=row_index,
                    start_col=col_index,
                    first_id=next_id,
                )
                items.extend(block_items)
                next_id += len(block_items)

        return items

    def _read_block(
        self,
        matrix: list[list],
        culture: str,
        prices: dict[tuple[str, str, str, str], float],
        start_row: int,
        start_col: int,
        first_id: int,
    ) -> list[dict]:
        location = cell_text(matrix, start_row + 3, start_col)
        packing = cell_text(matrix, start_row + 3, start_col + 2)
        updated_at = format_date(cell_value(matrix, start_row + 1, start_col))

        if not location or not packing:
            return []

        items = []
        row_index = start_row + 4
        product_order = 0
        current_product = ""
        current_product_order = -1
        while row_index < len(matrix):
            row_label = cell_text(matrix, row_index, start_col)
            if not row_label or normalize_text(row_label) == "total":
                break

            quality = quality_label(row_label)
            if quality and current_product:
                product = current_product
                order = current_product_order
                quality_order = QUALITY_STANDARDS.index(quality)
            else:
                product = row_label
                quality = ""
                order = product_order
                quality_order = -1
                current_product = product
                current_product_order = order
                product_order += 1

            quantity = parse_number(cell_value(matrix, row_index, start_col + 1))
            price = prices.get(price_key(culture, packing, product, quality), 0)
            if not price and not quality:
                price = prices.get(price_key(culture, packing, product, ""), 0)
            items.append(
                {
                    "id": first_id + len(items),
                    "local_carregamento": location,
                    "cultura": culture,
                    "packing": packing,
                    "produto": product,
                    "qualidade": quality,
                    "ordem_packing": start_row * 1000 + start_col,
                    "ordem_produto": order,
                    "ordem_qualidade": quality_order,
                    "quantidade": quantity,
                    "preco": price,
                    "atualizado_em": updated_at,
                }
            )
            row_index += 1

        fill_parent_quantities(items)
        return items


# ============================================================================
# Sincronizacao Git da planilha
# ============================================================================

class SpreadsheetGitSync:
    """Commita e envia ao Git somente alteracoes da planilha disponibilidade.xlsx."""

    def __init__(self, repo_dir: Path, spreadsheet_path: Path) -> None:
        self.repo_dir = repo_dir
        self.spreadsheet_path = spreadsheet_path
        self.relative_path = GIT_SPREADSHEET_PATH

    def sync_if_spreadsheet_changed(self) -> str:
        """Verifica mudanca na planilha e, quando existir, faz commit e push somente dela."""
        if not (self.repo_dir / ".git").exists() or not self.spreadsheet_path.exists():
            return ""

        if not self._acquire_lock():
            return "Sincronizacao Git da planilha ja esta em andamento."

        try:
            # A verificacao e limitada a disponibilidade.xlsx para impedir commit de codigo.
            status = self._git("status", "--porcelain", "--", self.relative_path)
            if not status.stdout.strip():
                return ""

            # Se ja existir outro arquivo staged, a automacao para para nao commitar codigo.
            staged_files = self._staged_files()
            if any(path != self.relative_path for path in staged_files):
                return "Sincronizacao Git pausada: ha outros arquivos preparados para commit."

            # O git add tambem recebe somente a planilha; outras mudancas ficam fora do commit.
            self._git("add", "--", self.relative_path)
            timestamp = datetime.now(BRASILIA_TZ).strftime("%d/%m/%Y %H:%M")
            commit = self._git(
                "commit",
                "-m",
                f"{GIT_COMMIT_MESSAGE} - {timestamp}",
                check=False,
            )
            if commit.returncode != 0:
                output = (commit.stderr or commit.stdout).strip()
                if "nothing to commit" in output.lower():
                    return ""
                return f"Falha ao commitar a planilha: {output}"

            push = self._git("push", "origin", self._current_branch(), check=False)
            if push.returncode != 0:
                output = (push.stderr or push.stdout).strip()
                return f"Commit da planilha criado, mas o push falhou: {output}"

            return "Planilha sincronizada automaticamente no Git."
        finally:
            self._release_lock()

    def _current_branch(self) -> str:
        branch = self._git("branch", "--show-current")
        return branch.stdout.strip() or "main"

    def _staged_files(self) -> list[str]:
        staged = self._git("diff", "--cached", "--name-only")
        return [line.strip() for line in staged.stdout.splitlines() if line.strip()]

    def _git(self, *args: str, check: bool = True) -> subprocess.CompletedProcess:
        # subprocess com lista de argumentos evita comandos compostos e limita a execucao ao Git.
        result = subprocess.run(
            ["git", "-C", str(self.repo_dir), *args],
            capture_output=True,
            text=True,
            timeout=45,
            check=False,
        )
        if check and result.returncode != 0:
            output = (result.stderr or result.stdout).strip()
            raise RuntimeError(output)
        return result

    def _acquire_lock(self) -> bool:
        # A trava evita que dois refreshes/sessoes tentem criar commit ao mesmo tempo.
        if GIT_SYNC_LOCK_PATH.exists():
            lock_age = datetime.now().timestamp() - GIT_SYNC_LOCK_PATH.stat().st_mtime
            if lock_age < GIT_SYNC_LOCK_MAX_AGE_SECONDS:
                return False

        GIT_SYNC_LOCK_PATH.write_text(
            datetime.now(BRASILIA_TZ).isoformat(),
            encoding="utf-8",
        )
        return True

    @staticmethod
    def _release_lock() -> None:
        try:
            GIT_SYNC_LOCK_PATH.unlink(missing_ok=True)
        except OSError:
            pass


# ============================================================================
# Estilo visual
# ============================================================================

class PageStyle:
    """Aplica CSS unico para as duas funcionalidades, com suporte real a dark mode."""

    @staticmethod
    def apply() -> None:
        st.set_page_config(
            page_title=APP_TITLE,
            page_icon="📦",
            layout="centered",
        )
        st.markdown(
            """
            <meta name="theme-color" content="#176b55">
            """,
            unsafe_allow_html=True,
        )

        st.markdown(
            """
            <style>

                :root {
                    color-scheme: light dark;

                    --bg: #f5f7fa;
                    --surface: #ffffff;
                    --surface-soft: #eef4ef;

                    --text: #17212b;
                    --text-soft: #4b5563;

                    --border: #d7dee7;

                    --green: #176b55;
                    --green-bright: #00a651;

                    --blue: #2f75b5;
                    --danger: #d32f2f;

                    --input-bg: #ffffff;
                    --input-text: #17212b;
                    --input-border: #cfd8e3;

                    --table-header: #176b55;
                }

                @media (prefers-color-scheme: dark) {

                    :root {

                        --bg: #0f1419;

                        --surface: #182028;

                        --surface-soft: #212b34;

                        --text: #f3f7fb;

                        --text-soft: #c5d0db;

                        --border: #344250;

                        --green: #1f8a67;

                        --green-bright: #37d487;

                        --blue: #4593d6;

                        --danger: #ff7b72;

                        --input-bg: #111827;

                        --input-text: #f3f7fb;

                        --input-border: #3b4a5a;

                        --table-header: #1f8a67;
                    }
                }

                html,
                body,
                .stApp {
                    background: var(--bg) !important;
                    color: var(--text) !important;
                }             
                .block-container {
                    max-width: 1040px;
                    padding-top: 2.4rem;
                    padding-bottom: 3rem;
                }

                /* =========================================
                   TEXTOS
                ========================================= */

                h1, h2, h3, h4, h5, h6,
                p, span, div, label,
                .stMarkdown,
                .stText,
                .stCaption {
                    color: var(--text) !important;
                }

                small {
                    color: var(--text-soft) !important;
                }

                /* =========================================
                   HEADER
                ========================================= */

                .app-header {
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    gap: 1.1rem;
                    margin-top: 0.4rem;
                    margin-bottom: 0.8rem;
                    text-align: center;
                }

                .app-logo {
                    width: 124px;
                    max-width: 32vw;
                    height: auto;
                    object-fit: contain;
                    display: block;
                }

                .app-title {
                    font-size: 2rem;
                    line-height: 1.1;
                    font-weight: 800;
                    color: var(--green-bright) !important;
                }

                /* =========================================
                   INPUTS STREAMLIT
                ========================================= */

                .stTextInput input,
                .stNumberInput input,
                .stSelectbox div[data-baseweb="select"],
                .stTextArea textarea {

                    background: var(--input-bg) !important;
                    color: var(--input-text) !important;

                    border: 1px solid var(--input-border) !important;

                    border-radius: 8px !important;
                }
                input,
                textarea,
                select {
                    font-size: 16px !important;
                }
                /* =========================================
                   SELECTBOX / DROPDOWN - DARK MODE REAL FIX
                    ========================================= */
                    
                    /* Caixa principal fechada */
                    
                    div[data-baseweb="select"] > div {
                        background: var(--input-bg) !important;
                        color: var(--input-text) !important;
                        border: 1px solid var(--input-border) !important;
                    }
                    
                    /* Texto dentro do select */
                    
                    div[data-baseweb="select"] span {
                        color: var(--input-text) !important;
                    }
                    
                    /* Ícone da seta */
                    
                    div[data-baseweb="select"] svg {
                    fill: var(--input-text) !important;
                    color: var(--input-text) !important;
                }
                    
                    /* Popup inteiro do dropdown */
                    
                    div[data-baseweb="popover"] {
                        background-color: var(--surface) !important;
                    }
                    div[data-baseweb="menu"] {
                        background: var(--surface) !important;
                        color: var(--text) !important;
                    }

                    /* Container da lista */
                    
                    ul[role="listbox"] {
                        background: var(--surface) !important;
                        border: 1px solid var(--border) !important;
                        border-radius: 10px !important;
                        padding: 4px !important;
                    }
                    
                    /* Cada opção */
                    
                    li[role="option"] {
                        background: var(--surface) !important;
                        color: var(--text) !important;
                        border-radius: 8px !important;
                    }
                    
                    /* Texto da opção */
                    
                    li[role="option"] * {
                        color: var(--text) !important;
                    }
                    
                    /* Hover */
                    
                    li[role="option"]:hover {
                        background: var(--surface-soft) !important;
                        color: var(--text) !important;
                    }
                    
                    /* Selecionado */
                    
                    li[aria-selected="true"] {
                        background: var(--green) !important;
                    }
                    
                    li[aria-selected="true"] * {
                        color: white !important;
                    }

                /* =========================================
              CHECKBOX - DARK MODE FIX REAL
                ========================================= */

                /* Texto */
                
                .stCheckbox label,
                .stCheckbox span {
                    color: var(--text) !important;
                    font-weight: 600;
                }
                
                /* Caixa externa */
                
                .stCheckbox div[role="checkbox"] {
                
                    background-color: var(--input-bg) !important;
                
                    border: 2px solid var(--input-border) !important;
                
                    border-radius: 6px !important;
                
                    width: 20px !important;
                
                    height: 20px !important;
                
                    transition: all 0.15s ease;
                }
                
                /* Hover */
                
                .stCheckbox div[role="checkbox"]:hover {
                
                    border-color: var(--green-bright) !important;
                
                    box-shadow: 0 0 0 1px var(--green-bright) !important;
                }
                
                /* Marcado */
                
                .stCheckbox div[role="checkbox"][aria-checked="true"] {
                
                    background-color: var(--green) !important;
                
                    border-color: var(--green) !important;
                }
                
                /* Ícone do check */
                
                .stCheckbox div[role="checkbox"] svg {
                
                    fill: white !important;
                
                    stroke: white !important;
                
                    stroke-width: 3 !important;
                    
                    color: white !important;
                
                    width: 16px !important;
                
                    height: 16px !important;
                }
                
                /* Remove fundo estranho do BaseWeb */
                
                .stCheckbox div[data-testid="stMarkdownContainer"] {
                
                    color: var(--text) !important;
                }
                
                /* Espaçamento melhor */
                
                .stCheckbox {
                
                    padding-top: 0.2rem;
                    padding-bottom: 0.2rem;
                }

                /* =========================================
                   BOTÕES
                ========================================= */

                .stButton button,
                .stDownloadButton button,
                .stFormSubmitButton button {

                    background: var(--green) !important;

                    color: white !important;

                    border: none !important;

                    border-radius: 8px !important;

                    font-weight: 700 !important;
                }

                .stButton button:hover,
                .stDownloadButton button:hover,
                .stFormSubmitButton button:hover {

                    filter: brightness(1.08);
                }

                /* =========================================
                   MÉTRICAS
                ========================================= */

                div[data-testid="stMetric"] {

                    border: 1px solid var(--border);

                    border-radius: 10px;

                    padding: 0.8rem;

                    background: var(--surface);
                }

                div[data-testid="stMetricLabel"] {

                    color: var(--text-soft) !important;
                }

                div[data-testid="stMetricValue"] {

                    color: var(--text) !important;

                    font-weight: 800;
                }

                /* =========================================
                   TABELAS
                ========================================= */

                .availability-table {

                    width: 100%;

                    border-collapse: collapse;

                    font-size: 0.92rem;

                    margin-bottom: 1rem;

                    background: var(--surface);

                    border: 1px solid var(--border);

                    border-radius: 8px;

                    overflow: hidden;
                }

                .availability-table th {

                    background: var(--table-header);

                    color: white;

                    font-weight: 800;

                    text-align: center;

                    padding: 0.45rem 0.5rem;

                    border: 1px solid var(--border);
                }

                .availability-table td {

                    color: var(--text);

                    text-align: center;

                    padding: 0.42rem 0.5rem;

                    border: 1px solid var(--border);
                }

                .availability-table tbody tr:nth-child(odd) td {

                    background: var(--surface-soft);
                }

                .availability-table tbody tr.product-row td {

                    font-weight: 800;

                    font-size: 0.95rem;
                }

                .availability-table tbody tr.product-row td:first-child {

                    text-align: left;

                    padding-left: 0.8rem;
                }

                .availability-table tbody tr.quality-row td:first-child {

                    padding-left: 1.7rem;

                    text-align: left;

                    color: var(--muted);
                }

                .availability-table tbody tr.quality-row td {

                    font-size: 0.82rem;

                    font-weight: 500;

                    color: var(--muted);
                }

                .availability-table tbody tr.total-row td {

                    background: var(--blue);

                    color: white;

                    font-weight: 800;
                }

                /* =========================================
                   CARDS DE COTAÇÃO
                ========================================= */

                .quote-result {

                    text-align: center;

                    padding: 0.9rem;

                    margin-top: 1rem;

                    border: 1px solid var(--border);

                    border-radius: 10px;

                    background: var(--surface);
                }

                .quote-price {

                    color: var(--danger);

                    font-size: 1.7rem;

                    font-weight: 800;

                    margin: 0.35rem 0;
                }

                .quote-title,
                .quote-section-title {

                    color: var(--green-bright) !important;
                }

                /* =========================================
                   ALERTAS
                ========================================= */

                .stAlert {
                    border-radius: 10px !important;
                }

                /* =========================================
                   FOOTER
                ========================================= */

                .app-footer {

                    margin-top: 2rem;

                    padding-top: 0.85rem;

                    border-top: 1px solid var(--border);

                    color: var(--text-soft);

                    text-align: center;

                    font-size: 0.85rem;
                }

                /* =========================================
                   FOCUS IOS/ANDROID
                ========================================= */

                input:focus,
                textarea:focus,
                select:focus {

                    outline: none !important;

                    border: 1px solid var(--green-bright) !important;

                    box-shadow: 0 0 0 1px var(--green-bright) !important;
                }

                /* =========================================
                   AUTOFILL ANDROID
                ========================================= */

                input:-webkit-autofill,
                input:-webkit-autofill:hover,
                input:-webkit-autofill:focus {

                    -webkit-text-fill-color: var(--input-text);

                    -webkit-box-shadow:
                        0 0 0px 1000px var(--input-bg) inset;

                    transition: background-color 9999s ease-in-out 0s;
                }

                /* =========================================
                   MOBILE
                ========================================= */

                @media (max-width: 700px) {

                    .app-header {
                        flex-direction: column;
                        gap: 0.6rem;
                    }

                    .app-logo {
                        width: 96px;
                    }

                    .app-title {
                        font-size: 1.65rem;
                    }

                    .quote-title {
                        font-size: 1.8rem;
                    }

                    .availability-table {
                        font-size: 0.8rem;
                    }

                    .availability-table th,
                    .availability-table td {
                        padding: 0.34rem;
                    }

                    .stButton button,
                    .stDownloadButton button,
                    .stFormSubmitButton button {

                        width: 100%;
                    }
                }

            </style>
            """,
            unsafe_allow_html=True,
        )


# ============================================================================
# Aplicacao Streamlit
# ============================================================================

class TrebeschiCommercialApp:
    """Coordena navegacao, tela de disponibilidade e tela de cotacao."""

    def __init__(self) -> None:
        self.catalog = AppCatalog()
        self.repository = ExcelAvailabilityRepository(
            excel_path=EXCEL_PATH,
            culture_names=self.catalog.culture_order,
        )

    def run(self) -> None:
        PageStyle.apply()
        self.disable_browser_translation()
        self.render_header()
        # O refresh global garante que a verificacao Git rode mesmo fora da tela de disponibilidade.
        self.enable_auto_refresh()
        self.render_spreadsheet_git_sync_status()
        selected_page = self.render_feature_selector()

        if selected_page == "Disponibilidade":
            self.render_availability_page()
        else:
            self.render_quote_page()

        self.render_footer()

    @staticmethod
    def disable_browser_translation() -> None:
        """Marca a pagina como pt-BR para evitar traducoes automaticas incorretas."""
        components.html(
            """
            <script>
                const doc = window.parent.document;
                doc.documentElement.setAttribute("lang", "pt-BR");
                doc.documentElement.setAttribute("translate", "no");
                doc.body.setAttribute("translate", "no");
            </script>
            """,
            height=0,
        )

    def render_header(self) -> None:
        logo_html = ""
        if LOGO_PATH.exists():
            logo_data = base64.b64encode(LOGO_PATH.read_bytes()).decode("ascii")
            logo_html = (
                f'<img class="app-logo" src="data:image/png;base64,{logo_data}" '
                'alt="Trebeschi">'
            )

        st.markdown(
            f"""
            <div class="app-header">
                {logo_html}
                <div class="app-title">{APP_TITLE}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    @staticmethod
    def render_feature_selector() -> str:
        st.markdown('<h4 class="feature-title">Escolha a funcionalidade</h4>', unsafe_allow_html=True)
        selected_page = st.radio(
            "Funcionalidade",
            ["Disponibilidade", "Cota\u00e7\u00e3o"],
            horizontal=True,
            label_visibility="collapsed",
        )
        st.divider()
        return selected_page

    def render_availability_page(self) -> None:
        st.caption(
            "Consulta de produtos disponiveis por local de carregamento."
        )
        self.render_update_status()
        self.render_consultation(self.repository.list_items())

    def render_quote_page(self) -> None:
        self.render_quote_header()
        self.render_quote_simulator()

    def render_update_status(self) -> None:
        if not EXCEL_PATH.exists():
            st.warning(f"Planilha disponibilidade.xlsx nao encontrada em: {EXCEL_PATH}")
            return

        modified_at = datetime.fromtimestamp(EXCEL_PATH.stat().st_mtime)
        st.caption(
            f"Ultima atualizacao da planilha: {modified_at.strftime('%d/%m/%Y %H:%M:%S')} "
            "| Atualizacao automatica a cada 10 minutos"
        )
        st.caption(f"Arquivo lido: {EXCEL_PATH}")

    @staticmethod
    def render_spreadsheet_git_sync_status() -> None:
        """Executa a automacao Git da planilha e exibe um retorno discreto para manutencao."""
        try:
            message = SpreadsheetGitSync(BASE_DIR, EXCEL_PATH).sync_if_spreadsheet_changed()
        except (RuntimeError, OSError, subprocess.SubprocessError) as error:
            st.warning(f"Sincronizacao Git da planilha nao concluida: {error}")
            return

        if message:
            # Sucesso fica visivel; pausas/falhas viram aviso para facilitar manutencao.
            normalized = normalize_text(message)
            if "falh" in normalized or "pausad" in normalized:
                st.warning(message)
            elif "sincronizada" in normalized:
                st.success(message)
            else:
                st.caption(message)

    @staticmethod
    def enable_auto_refresh() -> None:
        components.html(
            f"""
            <script>
                window.parent.document.documentElement.setAttribute("lang", "pt-BR");
                window.parent.document.body.setAttribute("translate", "no");
                setTimeout(function() {{
                    window.parent.location.reload();
                }}, {AUTO_REFRESH_SECONDS * 1000});
            </script>
            """,
            height=0,
        )

    def render_consultation(self, items: list[dict]) -> None:
        st.subheader("Disponibilidade por local")

        if not items:
            st.info("Nenhum produto encontrado na planilha.")
            return

        selected_location = st.selectbox(
            "Local de carregamento",
            self.catalog.location_names(),
            index=0,
        )
        available_cultures = self.catalog.cultures_for_location(selected_location, items)
        selected_culture_label = st.selectbox(
            "Cultura",
            ["Todas"] + [culture_label(culture) for culture in available_cultures],
        )
        selected_culture = culture_from_label(selected_culture_label)
        search = st.text_input(
            "Buscar produto ou packing",
            placeholder="Ex: AA Misto, Classe 5, LV",
        )

        filtered = self.filter_items(items, selected_location, selected_culture, search)
        filtered = [item for item in filtered if item["quantidade"] > 0]

        col1, col2 = st.columns(2)
        col1.metric("Itens", len(filtered))
        col2.metric("Qtd. disponivel", format_quantity(sum_quantities(filtered)) or "0")

        if not filtered:
            st.warning("Nenhum item encontrado com os filtros atuais.")
            return

        cultures_to_render = (
            [selected_culture] if selected_culture != "Todas" else available_cultures
        )
        for culture in cultures_to_render:
            culture_items = [item for item in filtered if item["cultura"] == culture]
            if culture_items:
                self.render_culture_table(culture, culture_items)

    def filter_items(
        self,
        items: list[dict],
        location: str,
        culture: str,
        search: str,
    ) -> list[dict]:
        filtered = list(items)

        if location != "Brasil":
            allowed_locations = self.catalog.excel_locations_for_filter(location)
            filtered = [
                item
                for item in filtered
                if item["local_carregamento"] in allowed_locations
            ]

        if culture != "Todas":
            filtered = [item for item in filtered if item["cultura"] == culture]

        if search:
            term = normalize_text(search)
            filtered = [
                item
                for item in filtered
                if term in normalize_text(item["produto"])
                or term in normalize_text(item["packing"])
                or term in normalize_text(item["local_carregamento"])
            ]

        return filtered

    def render_culture_table(self, culture: str, items: list[dict]) -> None:
        st.markdown(
            f'<div class="culture-title">{culture_label(culture)}</div>',
            unsafe_allow_html=True,
        )

        split_packings = unique_by_order(items, "packing", "ordem_packing")
        if len(split_packings) > 1:
            columns = st.columns(len(split_packings))
            for column, packing in zip(columns, split_packings):
                packing_items = [item for item in items if item["packing"] == packing]
                with column:
                    self.render_quantity_table(packing_title(culture, packing), packing_items)
            return

        self.render_quantity_table("", items)

    @staticmethod
    def render_quantity_table(title: str, items: list[dict]) -> None:
        if not items:
            return

        if title:
            st.markdown(f'<div class="table-subtitle">{title}</div>', unsafe_allow_html=True)

        columns = unique_quantity_columns(items)
        products = unique_by_order(items, "produto", "ordem_produto")
        rows = build_availability_rows(items, products, columns)
        st.markdown(render_html_table(rows), unsafe_allow_html=True)

    def render_quote_header(self) -> None:
        logo_html = ""
        if LOGO_PATH.exists():
            logo_data = base64.b64encode(LOGO_PATH.read_bytes()).decode("ascii")
            logo_html = (
                f'<img class="quote-logo" src="data:image/png;base64,{logo_data}" '
                'alt="Trebeschi">'
            )

    def render_quote_simulator(self) -> None:
        st.markdown(
            '<div class="quote-section-title">🧺 Selecione as culturas para cotação:</div>',
            unsafe_allow_html=True,
        )
        col_tomate, col_alho, col_cebola, col_batata_doce = st.columns(4)
        selected_cultures = {
            "Tomate": col_tomate.checkbox("🍅 Tomate", key="quote_tomate"),
            "Alho": col_alho.checkbox("🧄 Alho", key="quote_alho"),
            "Cebola": col_cebola.checkbox("🧅 Cebola", key="quote_cebola"),
            "Batata Doce": col_batata_doce.checkbox("🍠 Batata Doce", key="quote_batata_doce"),
        }

        if not any(selected_cultures.values()):
            st.info("☝️ Selecione pelo menos uma cultura para continuar.")
            return

        st.markdown(
            '<div class="quote-section-title">📋 Dados do cliente</div>',
            unsafe_allow_html=True,
        )
        quote_inputs = self.render_quote_form(selected_cultures)
        if quote_inputs is None:
            return

        quotes = calculate_quotes(**quote_inputs)
        validity = calculate_validity(quote_inputs["validity_input"])
        self.render_quote_results(quotes, validity)

    @staticmethod
    def render_quote_form(selected_cultures: dict[str, bool]) -> dict | None:
        with st.form("quote_form"):
            costs = {}
            if selected_cultures["Tomate"]:
                costs["Tomate"] = st.number_input("Custo Tomate (R$)", min_value=0.0, step=0.01)
            if selected_cultures["Alho"]:
                costs["Alho"] = st.number_input("Custo Alho (R$)", min_value=0.0, step=0.01)
            if selected_cultures["Cebola"]:
                costs["Cebola"] = st.number_input("Custo Cebola (R$)", min_value=0.0, step=0.01)
            if selected_cultures["Batata Doce"]:
                costs["Batata Doce"] = st.number_input("Custo Batata Doce (R$)", min_value=0.0, step=0.01)

            discount = st.number_input(
                "Desconto contratual / comissão (%)", min_value=0.0, step=0.01
            )
            logistics = st.number_input("Operação logística (R$)", min_value=0.0, step=0.01)
            tax_name = st.selectbox("Tipo de imposto", list(TAXES.keys()))

            tomato_package = None
            tomato_weight = None
            if selected_cultures["Tomate"]:
                st.markdown("#### Configuração Tomate")
                tomato_package = st.selectbox(
                    "Tipo de embalagem - Tomate",
                    ["Papelão", "HB", "IFCO", "Plástica"],
                )
                tomato_weight = st.selectbox(
                    "Peso por caixa - Tomate (kg)",
                    [18, 19, 20, 22, 23],
                )
                st.caption("Custo de mão de obra fixo: R$ 7,00")

            potato_package = None
            potato_weight = None
            if selected_cultures["Batata Doce"]:
                st.markdown("#### Configuração Batata Doce")
                potato_package = st.selectbox(
                    "Tipo de embalagem - Batata Doce",
                    ["Papelão", "IFCO", "HB", "SC"],
                )
                if potato_package == "SC":
                    potato_weight = 20
                    st.caption("Peso padrão para SC: 20 kg")
                else:
                    potato_weight = st.selectbox(
                        "Peso por caixa - Batata Doce (kg)",
                        [18, 20, 22,],
                    )

            validity_input = st.text_input(
                "Validade da cotação (hora e minuto, ex: 1430)",
                placeholder="ex: 1430",
            )
            submitted = st.form_submit_button("Calcular cotação", type="primary")

        if not submitted:
            st.info("Preencha os dados e clique em Calcular cotação.")
            return None

        return {
            "selected_cultures": selected_cultures,
            "costs": costs,
            "discount": discount,
            "logistics": logistics,
            "tax": TAXES[tax_name],
            "tomato_package": tomato_package,
            "tomato_weight": tomato_weight,
            "potato_package": potato_package,
            "potato_weight": potato_weight,
            "validity_input": validity_input,
        }

    def render_quote_results(self, quotes: list[dict], validity: datetime) -> None:
        data_str = validity.strftime("%d/%m/%Y")
        hora_str = validity.strftime("%H:%M")

        st.markdown(
            f"""
            <div class="quote-result">
                <h4 style="color: var(--danger); margin: 0.2rem 0;">COTAÇÃO TREBESCHI</h4>
                <strong>Válido até as {hora_str} hs do dia {data_str}.</strong><br>
                <span>Disponibilidade sujeita a alteração.</span>
            </div>
            """,
            unsafe_allow_html=True,
        )

        for quote in quotes:
            st.markdown(
                f"""
                <div class="quote-result">
                    <h3>{quote["icone"]} {quote["cultura"]}</h3>
                    <div>Embalagem: {quote["embalagem"]}</div>
                    <div class="quote-price">R$ {quote["valor_cx"]:.2f}</div>
                    <strong style="color: var(--green-bright);">
                        Valor por kg: R$ {quote["valor_kg"]:.2f}
                    </strong>
                </div>
                """,
                unsafe_allow_html=True,
            )

        try:
            pdf_buffer = generate_quote_pdf(quotes, LOGO_PATH, validity)
            st.download_button(
                label="Baixar PDF da cotação",
                data=pdf_buffer,
                file_name=f"Cotacao_Trebeschi_{data_str.replace('/', '-')}.pdf",
                mime="application/pdf",
            )
        except RuntimeError as error:
            st.error(str(error))

    @staticmethod
    def render_footer() -> None:
        st.markdown(
            """
            <div class="app-footer">
                Desenvolvido por Rafael Germano.<br>
                Para uso interno da Trebeschi.
            </div>
            """,
            unsafe_allow_html=True,
        )


# ============================================================================
# Calculos da cotacao
# ============================================================================

def calculate_quotes(
    selected_cultures: dict[str, bool],
    costs: dict[str, float],
    discount: float,
    logistics: float,
    tax: float,
    tomato_package: str | None,
    tomato_weight: int | None,
    potato_package: str | None,
    potato_weight: int | None,
    validity_input: str = "",
) -> list[dict]:
    """Calcula as cotacoes por cultura usando as mesmas premissas do app original."""

    del validity_input
    packing_labor = 7.00
    quotes = []

    if selected_cultures["Tomate"]:
        weight = tomato_weight or 20
        package = tomato_package or "Papelão"
        total_cost = ((costs["Tomate"] / 20) * weight) + logistics + packing_labor + PACKAGE_COSTS[package]
        box_value, kg_value = calculate_price(total_cost, weight, discount, tax)
        quotes.append(build_quote("Tomate", "🍅", f"{package} {weight} kg", box_value, kg_value))

    if selected_cultures["Alho"]:
        weight = 10
        total_cost = costs["Alho"] + (logistics * 0.6)
        box_value, kg_value = calculate_price(total_cost, weight, discount, tax)
        quotes.append(build_quote("Alho", "🧄", "Caixa 10 kg", box_value, kg_value))

    if selected_cultures["Cebola"]:
        weight = 20
        total_cost = costs["Cebola"] + logistics
        box_value, kg_value = calculate_price(total_cost, weight, discount, tax)
        quotes.append(build_quote("Cebola", "🧅", "Saco 20 kg", box_value, kg_value))

    if selected_cultures["Batata Doce"]:
        weight = potato_weight or 20
        package = potato_package or "Papelão"
        total_cost = ((costs["Batata Doce"] / 20) * weight) + logistics + PACKAGE_COSTS[package]
        box_value, kg_value = calculate_price(total_cost, weight, discount, tax)
        quotes.append(build_quote("Batata Doce", "🍠", f"{package} {weight} kg", box_value, kg_value))

    return quotes


def build_quote(culture: str, icon: str, package: str, box_value: float, kg_value: float) -> dict:
    return {
        "cultura": culture,
        "icone": icon,
        "embalagem": package,
        "valor_cx": box_value,
        "valor_kg": kg_value,
    }


def calculate_price(total_cost: float, weight: float, discount: float, tax: float) -> tuple[float, float]:
    box_value = total_cost / (1 - (discount / 100 + tax))
    return round_up_10_cents(box_value), box_value / weight


def round_up_10_cents(value: float) -> float:
    return ceil(value * 10) / 10


def calculate_validity(value: str) -> datetime:
    now = datetime.now(BRASILIA_TZ)
    if value and value.isdigit() and len(value) in [3, 4]:
        padded = value.zfill(4)
        try:
            hour = int(padded[:2])
            minute = int(padded[2:])
            validity = BRASILIA_TZ.localize(datetime.combine(date.today(), time(hour, minute)))
            if validity < now:
                validity += timedelta(days=1)
            return validity
        except ValueError:
            pass
    return now + timedelta(hours=2)


def generate_quote_pdf(quotes: list[dict], logo_path: Path, validity: datetime) -> BytesIO:
    """Gera PDF sob demanda; reportlab e importado aqui para o app abrir mesmo sem PDF."""

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfgen import canvas
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "A biblioteca reportlab nao esta instalada. Rode: pip install -r requirements.txt"
        ) from exc

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    if logo_path.exists():
        try:
            logo_img = PILImage.open(logo_path).convert("RGBA")
            background = PILImage.new("RGBA", logo_img.size, (255, 255, 255, 255))
            background.paste(logo_img, mask=logo_img)
            temp_logo = BytesIO()
            background.save(temp_logo, format="PNG")
            temp_logo.seek(0)
            pdf.drawImage(
                ImageReader(temp_logo),
                (width - 6 * cm) / 2,
                height - 6 * cm,
                width=6 * cm,
                height=3 * cm,
                mask="auto",
            )
        except OSError:
            pass

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, height - 7 * cm, "Cotacao Trebeschi")
    pdf.setFont("Helvetica", 12)
    pdf.drawCentredString(
        width / 2,
        height - 8 * cm,
        f"Valido ate as {validity.strftime('%H:%M')} hs do dia {validity.strftime('%d/%m/%Y')}.",
    )
    pdf.drawCentredString(width / 2, height - 9 * cm, "Disponibilidade sujeita a alteracao.")

    y = height - 10 * cm
    for quote in quotes:
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(3 * cm, y, quote["cultura"])
        y -= 0.6 * cm
        pdf.setFont("Helvetica", 11)
        pdf.drawString(3 * cm, y, f"Embalagem: {quote['embalagem']}")
        y -= 0.4 * cm
        pdf.drawString(3 * cm, y, f"Valor por embalagem: R$ {quote['valor_cx']:.2f}")
        y -= 0.4 * cm
        pdf.drawString(3 * cm, y, f"Valor por kg: R$ {quote['valor_kg']:.2f}")
        y -= 1.2 * cm

        if y < 5 * cm:
            pdf.showPage()
            y = height - 4 * cm

    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawCentredString(
        width / 2,
        3 * cm,
        "Gerado automaticamente pelo App Comercial Trebeschi",
    )
    pdf.save()
    buffer.seek(0)
    return buffer


# ============================================================================
# Utilitarios de tabela e texto
# ============================================================================

def build_availability_rows(items: list[dict], products: list[str], columns: list[str]) -> list[dict]:
    rows = []
    for product in products:
        row = {"Produto": product, "_row_type": "product", "_prices": {}}
        row_total = 0
        for column in columns:
            quantity = product_quantity(items, product, column)
            row[column] = quantity
            row["_prices"][column] = cell_price(items, product, column)
            row_total += quantity
        if row_total > 0:
            row["Total"] = row_total
            row["_prices"]["Total"] = unique_price_for_product(items, product)
            rows.append(row)

            for quality in unique_quality_labels(items, product):
                quality_row = {
                    "Produto": display_quality_label(quality),
                    "_row_type": "quality",
                    "_prices": {},
                }
                quality_total = 0
                for column in columns:
                    quantity = sum(
                        item["quantidade"]
                        for item in items
                        if item["produto"] == product
                        and item.get("qualidade") == quality
                        and quantity_column_name(item, items) == column
                    )
                    quality_row[column] = quantity
                    quality_row["_prices"][column] = cell_price(items, product, column, quality)
                    quality_total += quantity
                if quality_total > 0:
                    quality_row["Total"] = quality_total
                    quality_row["_prices"]["Total"] = unique_price_for_product(items, product, quality)
                    rows.append(quality_row)

    total_row = {"Produto": "TOTAL", "_row_type": "total"}
    for column in columns:
        total_row[column] = sum(product_quantity(items, product, column) for product in products)
    total_row["Total"] = sum(total_row[column] for column in columns)
    rows.append(total_row)
    return rows


def product_quantity(items: list[dict], product: str, column: str) -> float:
    parent_quantity = sum(
        item["quantidade"]
        for item in items
        if item["produto"] == product
        and not item.get("qualidade")
        and quantity_column_name(item, items) == column
    )
    if parent_quantity:
        return parent_quantity

    return sum(
        item["quantidade"]
        for item in items
        if item["produto"] == product
        and item.get("qualidade")
        and quantity_column_name(item, items) == column
    )


def cell_price(items: list[dict], product: str, column: str, quality: str = "") -> float:
    prices = {
        item.get("preco", 0)
        for item in items
        if item["produto"] == product
        and item.get("qualidade", "") == quality
        and quantity_column_name(item, items) == column
        and item.get("preco", 0) > 0
    }
    return prices.pop() if len(prices) == 1 else 0


def unique_price_for_product(items: list[dict], product: str, quality: str = "") -> float:
    prices = {
        item.get("preco", 0)
        for item in items
        if item["produto"] == product
        and item.get("qualidade", "") == quality
        and item.get("preco", 0) > 0
    }
    return prices.pop() if len(prices) == 1 else 0


def unique_quality_labels(items: list[dict], product: str) -> list[str]:
    quality_items = [
        item
        for item in items
        if item.get("qualidade") and (not product or item["produto"] == product)
    ]
    quality_items.sort(key=lambda item: item.get("ordem_qualidade", 0))
    labels = []
    for item in quality_items:
        quality = item["qualidade"]
        if quality not in labels:
            labels.append(quality)
    return labels


def fill_parent_quantities(items: list[dict]) -> None:
    for item in items:
        if item.get("qualidade") or item["quantidade"]:
            continue

        item["quantidade"] = sum(
            child["quantidade"]
            for child in items
            if child["produto"] == item["produto"]
            and child.get("qualidade")
            and child["local_carregamento"] == item["local_carregamento"]
            and child["packing"] == item["packing"]
        )


def cell_value(matrix: list[list], row: int, col: int):
    if row >= len(matrix) or col >= len(matrix[row]):
        return None
    return matrix[row][col]


def cell_from_row(row: tuple, index: int):
    if index >= len(row):
        return None
    return row[index]


def cell_text(matrix: list[list], row: int, col: int) -> str:
    value = cell_value(matrix, row, col)
    return "" if value is None else str(value).strip()


def parse_number(value) -> float:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(".", "").replace(",", "."))
    except ValueError:
        return 0


def price_key(culture: str, packing: str, product: str, quality: str = "") -> tuple[str, str, str, str]:
    return (
        normalize_text(culture),
        normalize_text(packing),
        normalize_text(product),
        normalize_text(short_quality_label(quality)),
    )


def format_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if value:
        return str(value)
    return datetime.now().strftime("%d/%m/%Y")


def normalize_text(value) -> str:
    text = "" if value is None else str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(char for char in text if not unicodedata.combining(char))


def quality_label(value) -> str:
    normalized = normalize_text(value)
    for quality in QUALITY_STANDARDS:
        if normalized == normalize_text(quality):
            return quality
    return ""


def short_quality_label(value) -> str:
    text = "" if value is None else str(value)
    return text.replace("Padrão ", "").replace("Padrao ", "")


def display_quality_label(value) -> str:
    text = "" if value is None else str(value).strip()
    if normalize_text(text) == "a":
        return "Padrão A"
    if normalize_text(text) == "b":
        return "Padrão B"
    return text


def culture_label(culture: str) -> str:
    if culture == "Todas":
        return culture
    return f"{CULTURE_ICONS.get(culture, '•')} {culture}"


def culture_from_label(label: str) -> str:
    if label == "Todas":
        return "Todas"
    return label.split(" ", 1)[1] if " " in label else label


def packing_title(culture: str, packing: str) -> str:
    tomato_titles = {
        "lv": "Longa Vida (LV)",
        "it": "Italiano (IT)",
    }
    if culture == "Tomate":
        return tomato_titles.get(normalize_text(packing), packing)
    return packing


def quantity_column_name(item: dict, items: list[dict]) -> str:
    locations = {row["local_carregamento"] for row in items}
    if len(locations) == 1:
        return item["packing"]
    return f"{item['local_carregamento']} - {item['packing']}"


def format_quantity(value: float) -> str:
    if not value:
        return ""
    return f"{value:,.0f}".replace(",", ".")


def format_price(value: float) -> str:
    if not value:
        return ""
    formatted = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {formatted}"


def format_quantity_with_price(quantity: float, price: float) -> str:
    quantity_text = format_quantity(quantity)
    price_text = format_price(price)
    if quantity_text and price_text:
        return f"{quantity_text} ({price_text})"
    return quantity_text


def render_html_table(rows: list[dict]) -> str:
    if not rows:
        return ""

    headers = [header for header in rows[0].keys() if not header.startswith("_")]
    header_html = "".join(
        f'<th translate="no">{escape_html(header)}</th>' for header in headers
    )
    body_html = []

    for row in rows:
        row_type = row.get("_row_type", "")
        row_class = f' class="{row_type}-row"' if row_type else ""
        prices = row.get("_prices", {})
        cells = []
        for header in headers:
            value = row.get(header, "")
            if header != "Produto":
                value = format_quantity_with_price(value, prices.get(header, 0))
            cells.append(f'<td translate="no">{escape_html(value)}</td>')
        body_html.append(f"<tr{row_class}>{''.join(cells)}</tr>")

    return (
        '<table class="availability-table" translate="no">'
        f"<thead><tr>{header_html}</tr></thead>"
        f"<tbody>{''.join(body_html)}</tbody>"
        "</table>"
    )


def escape_html(value) -> str:
    text = "" if value is None else str(value)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )


def sum_quantities(items: list[dict]) -> float:
    return sum(item["quantidade"] for item in items)


def unique_by_order(items: list[dict], value_key: str, order_key: str) -> list:
    ordered_values = {}
    for item in items:
        value = item[value_key]
        order = item.get(order_key, 0)
        if value not in ordered_values or order < ordered_values[value]:
            ordered_values[value] = order
    return [
        value
        for value, _ in sorted(
            ordered_values.items(),
            key=lambda pair: (pair[1], str(pair[0])),
        )
    ]


def unique_quantity_columns(items: list[dict]) -> list:
    ordered_columns = {}
    for item in items:
        column = quantity_column_name(item, items)
        order = item.get("ordem_packing", 0)
        if column not in ordered_columns or order < ordered_columns[column]:
            ordered_columns[column] = order
    return [
        column
        for column, _ in sorted(
            ordered_columns.items(),
            key=lambda pair: (pair[1], str(pair[0])),
        )
    ]


if __name__ == "__main__":
    TrebeschiCommercialApp().run()
